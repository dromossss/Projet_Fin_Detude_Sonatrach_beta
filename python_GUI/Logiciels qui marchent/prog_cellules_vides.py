import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import openpyxl
from openpyxl.styles import PatternFill

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        
        workbook = openpyxl.load_workbook(file_path) #Chaaarger le fichier choisi
        sheet = workbook.active

        #Iteratation dans toutes les cellules de la feuille
        for row in sheet.iter_rows():
            for cell in row:
                # vérifier si la cellule = none
                if cell.value is None:
                    #Highlight des cellules vides
                    
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") #couleur choisi = yellow

        
        workbook.save(file_path) #sauvegarde du fichier modifié
        print("Cellules vides mises en évidence et le fichier excel a été sauvegardé.") 

        #ouvrire le fichier excel modifié
        open_file(file_path)

def open_file(file_path): 
    #associé le systeme par defaut avec les fichiers xlsx
    import os #importer le systeme
    os.startfile(file_path)


root = tk.Tk() #Creer fenetre tkinter
root.title("Highlighting des cellules vides") #nom de la fenetre
root.geometry("500x250") #taille

#appliquer style forest dark
root.tk.call("source", r"C:\Users\user\Documents\PYTHON\python_GUI\Logiciels qui marchent\Forest-ttk-theme-master\Forest-ttk-theme-master\forest-dark.tcl")


style = ttk.Style(root) #configuration du style
style.theme_use("forest-dark")
style.configure(".", font=("Arial", 10))


select_file_btn = ttk.Button(root, text="Selection du fichier Excel", command=select_file) #Bouton de selection
select_file_btn.pack(pady=80)

root.mainloop()
