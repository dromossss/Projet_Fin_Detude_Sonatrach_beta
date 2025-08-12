import tkinter as tk
from tkinter import ttk
import openpyxl

def load_data():
    path = r"C:\Users\ryadk\OneDrive\Documents\flash red\APPLICATION\python_GUI\Logiciels qui marchent\Forest-ttk-theme-master\Forest-ttk-theme-master\forest-dark.tcl"
    workbook = openpyxl.load_workbook(path)
    data = workbook.active
    
    list_values = list(data.values)
    print(list_values)
    for col_name in list_values[0]:
        treeview.heading(col_name, text=col_name)
        
    for value_tuple in list_values[1:]:
        treeview.insert('', tk.END, values=value_tuple)
        
def show_message(message):
    message_window = tk.Toplevel(root)
    message_window.title("Message")

    message_label = tk.Label(message_window, text=message)
    message_label.pack()

    ok_button = tk.Button(message_window, text="OK", command=message_window.destroy)
    ok_button.pack()

def insert_row():
    # vérifier si une entrée est vide
    if not MATRICULE_entry.get() or not nom_entry.get() or not CodeDom_entry.get() or not Cout_entry.get():
        show_message("Un ou plusieurs champs ont été oubliés")
        return

    # Inserer les lignes dans data
    MATRICULE = MATRICULE_entry.get()       # Matricule
    NOM = nom_entry.get()                   # Nom prénom
    genre = status_combobox.get()                 # Genre
    rang = status_combobox1.get()                  # Rang

    CodeDomaine = CodeDom_entry.get()       # Code domaine
    choix_domaine = status_combobox2.get()        # Choix domaine
    mode_status = "E-Learning" if a.get() else "Présentiel"  # Choix mode

    Cout = Cout_entry.get()                 # Cout globale

    # Inserer données dans 'data'
    path = r"C:\Users\user\Documents\PYTHON\python_GUI\Logiciels qui marchent\programme_insertion\ex_insertion.xlsx"
    workbook = openpyxl.load_workbook(path)
    data = workbook.active
    last_row = data.max_row + 1
    data.cell(row=last_row, column=1).value = MATRICULE
    data.cell(row=last_row, column=2).value = NOM
    data.cell(row=last_row, column=3).value = genre
    data.cell(row=last_row, column=4).value = rang
    data.cell(row=last_row, column=5).value = CodeDomaine
    data.cell(row=last_row, column=6).value = choix_domaine
    data.cell(row=last_row, column=7).value = mode_status
    data.cell(row=last_row, column=8).value = Cout
    workbook.save(path)

    #Inserer dans le treeview
    treeview.insert("", tk.END, values=(MATRICULE, NOM, genre, rang, CodeDomaine, choix_domaine, mode_status, Cout))

root = tk.Tk()

#Appliquer le theme
style = ttk.Style(root)
root.call("source", r"C:\Users\ryadk\OneDrive\Documents\flash red\APPLICATION\python_GUI\Logiciels qui marchent\Forest-ttk-theme-master\Forest-ttk-theme-master\forest-dark.tcl")
root.call("source", r"C:\Users\ryadk\OneDrive\Documents\flash red\APPLICATION\python_GUI\Logiciels qui marchent\Forest-ttk-theme-master\Forest-ttk-theme-master\forest-light.tcl")

#choisir theme
style.theme_use("forest-dark")

#combobox liste:
combo_list = ["Non Précisé" , "Male" , "Femelle"]
combo_list1 = ["C" , "E" , "M"]
combo_list2 = ["Non Précisé" , "Etudes" , "Juridique" , "Communication", "Maintenance" , "Management", "RH" , "Sécurité Ind", "Controle Qualité"]

#creer fenetre (frame)
frame = ttk.Frame(root)
frame.pack()

#nom et widgets fenetre
widgets_frame = ttk.LabelFrame(frame, text="Infos Personelles")
widgets_frame.grid(row=0, column=0, padx=10, pady=(5))

#fenetres d'insertion
MATRICULE_entry = ttk.Entry(widgets_frame)
MATRICULE_entry.insert(0,"Matricule")
MATRICULE_entry.bind("<FocusIn>", lambda e: MATRICULE_entry.delete('0','end'))
MATRICULE_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

nom_entry = ttk.Entry(widgets_frame)
nom_entry.insert(0,"Nom et Prénom")
nom_entry.bind("<FocusIn>", lambda e: nom_entry.delete('0','end'))
nom_entry.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="ew")

#creer drop down menu

#frame genre
widgets_frame = ttk.LabelFrame(frame, text="Genre")
widgets_frame.grid(row=1, column=0, padx=20, pady=10)
#choisir genre utilisateur
status_combobox = ttk.Combobox(widgets_frame, values=combo_list)
status_combobox.current(0) 
status_combobox.grid(row=2, column=0, padx=5, pady=(0, 5), sticky="ew")

#frame rang
widgets_frame = ttk.LabelFrame(frame, text="Rang")
widgets_frame.grid(row=3, column=0, padx=20, pady=10)
#choisir Rang
status_combobox1 = ttk.Combobox(widgets_frame, values=combo_list1)
status_combobox1.current(0) 
status_combobox1.grid(row=3, column=0, padx=5, pady=(0, 5), sticky="ew")

#frame domaine
widgets_frame = ttk.LabelFrame(frame, text="Domaine")
widgets_frame.grid(row=4, column=0, padx=20, pady=10)

#widget entry code domaine
CodeDom_entry = ttk.Entry(widgets_frame)
CodeDom_entry.insert(0,"Code Domaine")
CodeDom_entry.bind("<FocusIn>", lambda e: CodeDom_entry.delete('0','end'))
CodeDom_entry.grid(row=3, column=0, padx=5, pady=(0, 5), sticky="ew")

#choix domaine
status_combobox2 = ttk.Combobox(widgets_frame, values=combo_list2)
status_combobox2.current(0) 
status_combobox2.grid(row=4, column=0, padx=5, pady=(0, 5), sticky="ew")

#widget entry cout globale
Cout_entry = ttk.Entry(widgets_frame)
Cout_entry.insert(0,"Couts")
Cout_entry.bind("<FocusIn>", lambda e: Cout_entry.delete('0','end'))
Cout_entry.grid(row=5, column=0, padx=5, pady=(0, 5), sticky="ew")

#checkbox pour presentiel ou e-learning

#boucle pour selectionner une des deux seulement
def handle_check():
    if a.get() == 1:
        b.set(0)
    elif b.get() == 1:
        a.set(0)

def show_error_message():
    error_window = tk.Toplevel(root)
    error_window.title("Erreur")
    error_label = tk.Label(error_window, text="Un ou plusieurs champs ont été oubliés")
    error_label.pack(padx=10, pady=10)
    ok_button = ttk.Button(error_window, text="OK", command=error_window.destroy)
    ok_button.pack(padx=10, pady=10)

#configuration boutons
a = tk.BooleanVar()
checkbutton_a = ttk.Checkbutton(widgets_frame, text="E-Learning", variable=a, command=handle_check)
checkbutton_a.grid(row=6, column=0, padx=5, pady=(0, 5), sticky="nsew")

b = tk.BooleanVar()
checkbutton_b = ttk.Checkbutton(widgets_frame, text="Présentiel", variable=b, command=handle_check)
checkbutton_b.grid(row=6, column=1, padx=5, pady=(0, 5), sticky="nsew")

#creer bouton "insérer"
button = ttk.Button(widgets_frame, text="Insérer", command=insert_row)
button.grid(row= 7, column=0, padx=5, pady=(0, 5), sticky="nsew")

treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10)

treeScrollY = ttk.Scrollbar(treeFrame)
treeScrollY.pack(side="right", fill="y")

treeScrollX = ttk.Scrollbar(treeFrame, orient="horizontal")
treeScrollX.pack(side="bottom", fill="x")

cols = ("MATRICULE", "NOM PRENOM", "GENRE", "RANG", "CODE DOMAINE", "DOMAINE", "MODE FORMATION", "COUT GLOBALE")
treeview = ttk.Treeview(treeFrame, show="headings", yscrollcommand=treeScrollY.set, xscrollcommand=treeScrollX.set, columns=cols, height=10)

treeview.column("MATRICULE", width=100)
treeview.column("NOM PRENOM", width=100)
treeview.column("GENRE", width=50)
treeview.column("RANG", width=50)
treeview.column("CODE DOMAINE", width=100)
treeview.column("DOMAINE", width=100)
treeview.column("MODE FORMATION", width=100)
treeview.column("COUT GLOBALE", width=100)


treeview.pack()
treeScrollY.config(command=treeview.yview)
treeScrollX.config(command=treeview.xview)


load_data()

root.mainloop()
