import tkinter as tk
from tkinter import filedialog
import openpyxl
import os

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        return file_path

def calculate_sum():
    file_path = select_file()
    if file_path:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        column_window = tk.Toplevel(root)
        column_window.title("Selectionner colonne")

        column_var = tk.StringVar()
        column_label = tk.Label(column_window, text="Selectionner colonne (A, B, C...):")
        column_label.pack()
        column_entry = tk.Entry(column_window, textvariable=column_var)
        column_entry.pack()
        calculate_sum_btn = tk.Button(column_window, text="Calculer somme", command=lambda: handle_calculate_sum(workbook, sheet, file_path, column_var.get()))
        calculate_sum_btn.pack()

def handle_calculate_sum(workbook, sheet, file_path, selected_column):
    if add_values(workbook, sheet, file_path, selected_column):
        show_message("Opération réussie")
        open_excel_file(file_path)
    else:
        show_message("Valeur invalide, réessayez")

def add_values(workbook, sheet, file_path, selected_column):
    column_values = [cell.value for cell in sheet[selected_column][1:] if cell.value is not None and isinstance(cell.value, (int, float))]

    total_sum = sum(column_values)# Calculer la somme

    last_row_index = len(sheet[selected_column]) #index de la derniere ligne

    sum_cell = sheet.cell(row=last_row_index+2, column=sheet[selected_column][0].column)  # Creer cellules pour la somme et pour "Total Globale"
    sum_cell.value = total_sum

    total_cell = sheet.cell(row=last_row_index+3, column=sheet[selected_column][0].column)
    total_cell.value = "Total Globale"

    workbook.save(file_path)
    return True

def open_file(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(values_only=True):
        print(*row, sep="\t")

def show_message(message):
    message_window = tk.Toplevel(root)
    message_window.title("Message")
    message_label = tk.Label(message_window, text=message)
    message_label.pack()
    ok_button = tk.Button(message_window, text="OK", command=message_window.destroy)
    ok_button.pack()

def open_excel_file(file_path):
    os.system(f'start excel "{file_path}"')

root = tk.Tk() #creer fenetre tkinter
root.title("Calculer le Totale Global") #donner un nom a la fenetre
root.geometry("300x150")

root.tk.call("source", r"C:\Users\user\Documents\PYTHON\python_GUI\Logiciels qui marchent\Forest-ttk-theme-master\Forest-ttk-theme-master\forest-dark.tcl")

#bouton calculer somme
calculate_sum_btn = tk.Button(root, text="Calculer la somme", command=calculate_sum)
calculate_sum_btn.pack(pady=50)

root.mainloop()
