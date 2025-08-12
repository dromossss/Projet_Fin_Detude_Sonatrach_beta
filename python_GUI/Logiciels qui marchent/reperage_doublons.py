import tkinter as tk
from tkinter import filedialog
import openpyxl
from openpyxl.styles import PatternFill
from collections import Counter
import os

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        return file_path

def trouver_doublons():
    file_path = select_file()
    if file_path:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        column_window = tk.Toplevel(root)
        column_window.title("Selection colonne")

        def handle_trouver_doublons():
            selected_column = column_var.get()
            if selected_column.isalpha():
                if highlight_doublons(workbook, sheet, file_path, selected_column):
                    show_message("Opération réussie")
                    column_window.destroy()
                    open_excel_file(file_path)
                else:
                    show_message("Valeur entrée incorrecte, réessayez")
            else:
                show_message("Valeur entrée incorrecte, réessayez")

        column_var = tk.StringVar()
        column_label = tk.Label(column_window, text="Selection colonne (A, B, C...):")
        column_label.pack()
        column_entry = tk.Entry(column_window, textvariable=column_var)
        column_entry.pack()
        trouver_doublons_btn = tk.Button(column_window, text="OK", command=handle_trouver_doublons)
        trouver_doublons_btn.pack()

def highlight_doublons(workbook, sheet, file_path, selected_column):
    column_values = [cell.value for cell in sheet[selected_column] if cell.value is not None]

    doublons = [item for item, count in Counter(column_values).items() if count > 1]

    colors = ["FF0000", "00FF00", "0000FF", "FFFF00", "FF00FF", "00FFFF"]
    color_mapping = {}
    color_index = 0

    for cell in sheet[selected_column]:
        if cell.value is not None and cell.value in doublons:
            if cell.value not in color_mapping:
                color_mapping[cell.value] = colors[color_index % len(colors)]
                color_index += 1
            cell.fill = PatternFill(start_color=color_mapping[cell.value], end_color=color_mapping[cell.value], fill_type="solid")

    workbook.save(file_path)
    print("Doublons trouvés et fichier sauvegardé")
    return True

def show_message(message):
    message_window = tk.Toplevel(root)
    message_window.title("Message")
    message_label = tk.Label(message_window, text=message)
    message_label.pack()
    ok_button = tk.Button(message_window, text="OK", command=message_window.destroy)
    ok_button.pack()

def open_excel_file(file_path):
    os.startfile(file_path)

#fenetre tkinter
root = tk.Tk()
root.title("Repérage doublons")
root.geometry("300x150")

#selection du theme
root.tk.call("source", r"C:\Users\user\Documents\PYTHON\python_GUI\Logiciels qui marchent\Forest-ttk-theme-master\Forest-ttk-theme-master\forest-dark.tcl")

#bouton :trouver_doublons
trouver_doublons_btn = tk.Button(root, text="Trouver doublons", command=trouver_doublons)
trouver_doublons_btn.pack(pady=50)

root.mainloop()
