import openpyxl
from openpyxl.styles import Font

wb = openpyxl.load_workbook(r"C:\Users\user\Documents\PYTHON\python bits\balance.xlsx")
ws = wb['data']

#une seule cellule, changer font, color ect...
#font_style = Font(name="Chalkboard", size=14, color="1A4FDF", italic=True, bold=True)
#a4 = ws['A4']
#a4.font = font_style  # Use a4.font instead of a4.style
#wb.save(r"C:\Users\user\Documents\PYTHON\python bits\balance.xlsx")

#modifier font de plusieurs cellules d'une colonne:
#col_style = Font(name="Reem Kufi", size=12, color="DB3B22", underline='single', strikethrough=True)

#for i in range (2,10):
#    ws.cell(row=i, column=3).font = col_style
    
#wb.save(r"C:\Users\user\Documents\PYTHON\python bits\balance.xlsx")