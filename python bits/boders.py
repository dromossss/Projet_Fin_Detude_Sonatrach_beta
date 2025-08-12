import openpyxl
from openpyxl.styles import Border , Side

wb = openpyxl.load_workbook(r"C:\Users\user\Documents\PYTHON\python bits\balance.xlsx")
ws= wb['data']

top = Side(border_style='dashed', color ="FF0707")
bottom = Side(border_style='double', color ="10AF2A")

#methode 1:
border = Border(top=top)
ws['B6'].border = border

border = Border(bottom=bottom)
ws['B7'].border = border

#methode 2 :
border = Border(top=top, bottom=bottom)
ws['B6:B8'].border = border

wb.save(r"C:\Users\user\Documents\PYTHON\python bits\balance.xlsx")