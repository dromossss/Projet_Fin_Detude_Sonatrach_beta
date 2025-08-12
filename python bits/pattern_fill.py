import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook(r"C:\Users\user\Documents\PYTHON\python bits\balance.xlsx") 
ws = wb['data']

#fill pattern donc colorier la cellule
fill_pattern = PatternFill(patternType='solid', fgColor='C64747')

ws['B4'].fill = fill_pattern
wb.save(r"C:\Users\user\Documents\PYTHON\python bits\balance.xlsx")

#row 9 = 13+2 = 15
# row 9 reward : 1604 + 14 = 1618
#row 5 = 68 + 22 = 90
# row 5 = 6937 + 7 = 6944