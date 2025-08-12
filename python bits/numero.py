import openpyxl
from openpyxl.styles import numbers

wb = openpyxl.load_workbook(r"C:\Users\user\Documents\PYTHON\python bits\balance.xlsx")
ws= wb['data']

ws['D4'] = '11/11/20'
ws['D4'].number_format = numbers.FORMAT_DATE_DATETIME


ws['E4'] = 20

ws['F4'] = 'Beginner'
ws['F4'].number_format = numbers.format_text

wb.save(r"C:\Users\user\Documents\PYTHON\python bits\balance.xlsx")

