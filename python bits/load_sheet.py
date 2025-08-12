import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\user\Documents\PYTHON\python bits\balance.xlsx")
ws = wb['data']

#ws['C1']='Double Balance'
#wb.save(r"C:\Users\user\Documents\PYTHON\python bits\balance.xlsx")

#boucle pour cloner une colonne et la multiplier par 2:
#for i in range (2,10):
 #   b_col = ws.cell(row=i, column=2).value
  #  c_value = b_col * 2
   # ws.cell(row=i, column=3).value = c_value
    #wb.save(r"C:\Users\user\Documents\PYTHON\python bits\balance.xlsx")

#changer une valeur d'une cellule:
#ws.cell(row=5, column= 2).value = 50
#wb.save(r"C:\Users\user\Documents\PYTHON\python bits\balance.xlsx")


#lire valeurs dans une cell precise:
#print (ws.cell(row=5, column= 2).value)


#inserer valeurs dans des cases précises:
#ws['A9'] = 'Raymond'
#ws['B9'] = 16965
#wb.save(r"C:\Users\user\Documents\PYTHON\python bits\balance.xlsx")


#ws['C1']='Double Balance'
#wb.save(r"C:\Users\user\Documents\PYTHON\python bits\balance.xlsx")

#ws.cell(row=5, column=2).value = 80000
#ws.cell(row=5, column=2)
#wb.save(r"C:\Users\user\Documents\PYTHON\python bits\balance.xlsx")

#print(ws['B8'].value)

#ajouter a une case et sauvegarder
#ws['B5'].value =9
#wb.save(r"C:\Users\user\Documents\PYTHON\python bits\balance.xlsx")

#rows = ws.iter_rows(min_row=1, max_row=7, min_col=1, max_col=2)
#print (rows)

#names = []
#balance = []
#for a,b in rows:
    #names.append(a.value)
   # balance.append(b.value)
    
    #print(names)
    #print(balance)
    
#columns = ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=2)
#for col in columns :
    #print(col)

#value_range = ws['A2':'B5']

#print(value_range)

#for a, b in value_range:
 #   print(a.value, b.value)



#print(ws['a5'].value)

#print(wb.sheetnames)

#ws= wb['data']
#print(ws)

#ws1= wb['Sheet1']
#print(ws1)

#wb.create_sheet("New_Sheet3", 3)

#wb.save(r"C:\Users\user\Documents\PYTHON\python bits\balance.xlsx")


