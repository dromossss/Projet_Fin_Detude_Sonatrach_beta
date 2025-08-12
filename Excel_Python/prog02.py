import openpyxl
import tkinter as tk
from tkinter import messagebox

fichiers = [
    'C:\\Users\\user\\Documents\\PYTHON\\Excel_Python\\excel01\\PMT_ACTION.xlsx',
    'C:\\Users\\user\\Documents\\PYTHON\\Excel_Python\\excel01\\PMT_PAR_CSP.xlsx',
    'C:\\Users\\user\\Documents\\PYTHON\\Excel_Python\\excel01\\PMT_PAR_TYPE_DE_FORMATION.xlsx'
]

valeurs = []

# Fonction pour collecter le contenu d'une cellule précise à partir d'un fichier Excel
def collect_cell_content(file_path, sheet_name, cell):
    wb = openpyxl.load_workbook(file_path)
    feuille = wb[sheet_name]
    valeur = feuille[cell].value
    wb.close()  # Fermer le classeur après avoir récupéré la valeur
    return valeur

# Fonction pour enregistrer les modifications dans les fichiers Excel
def sauvegarder_modifications():
    for fichier in fichiers:
        wb = openpyxl.load_workbook(fichier)
        wb.save(fichier)
        wb.close()  # Fermer chaque classeur après l'enregistrement des modifications

# Fonction pour gérer le clic sur le bouton Ajouter
def ajouter_valeur():
    # Récupérer la cellule et la valeur entrées par l'utilisateur
    cellule = entry_cellule.get()
    donnee = entry_valeur.get()

    # Parcourir les fichiers et ajouter la donnée dans la cellule choisie dans chaque fichier
    for fichier in fichiers:
        wb = openpyxl.load_workbook(fichier)
        feuille = wb['Sheet1']
        feuille[cellule] = donnee
        wb.close()  # Fermer chaque classeur après l'ajout de la donnée

    # Collecter le contenu de la cellule choisie dans chaque fichier
    valeurs.clear()
    for fichier in fichiers:
        valeur = collect_cell_content(fichier, 'Sheet1', cellule)
        valeurs.append(valeur)

    # Afficher la liste des valeurs collectées
    messagebox.showinfo("Valeurs Collectées", str(valeurs))

    # Appeler la fonction pour sauvegarder les modifications dans les fichiers Excel
    sauvegarder_modifications()

# Créer la fenêtre principale
root = tk.Tk()
root.title("Ajouter une Valeur")

# Créer les labels et les entrys pour la cellule et la valeur
label_cellule = tk.Label(root, text="Cellule (Alphabet, Number):")
label_cellule.pack()
entry_cellule = tk.Entry(root)
entry_cellule.pack()

label_valeur = tk.Label(root, text="Valeur:")
label_valeur.pack()
entry_valeur = tk.Entry(root)
entry_valeur.pack()

# Créer le bouton Ajouter
button_ajouter = tk.Button(root, text="Ajouter", command=ajouter_valeur)
button_ajouter.pack()

# Lancer la boucle principale de l'interface graphique
root.mainloop()
